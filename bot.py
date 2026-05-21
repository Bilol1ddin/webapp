"""
O'quv Markaz Bot — v4 (To'liq)
================================
Rollar:
  ADMIN    — hamma narsa
  TEACHER  — faqat o'z guruhining davomati
  STUDENT  — dars jadvali, davomat ko'rish, eslatma

Tuzatilgan:
  • Davomat belgilash — guruh callback ishlaydi
  • guruhlar_royxati — davomat.json bo'lmasa ham ko'rsatadi
  • O'qituvchi roli (TEACHERS sozlamada)

SOZLASH:
  BOT_TOKEN, ADMIN_IDS, TEACHERS, JADVAL_VARAQLAR
  pip install "python-telegram-bot[job-queue]" openpyxl
  python bot.py
"""

import calendar, difflib, json, logging, random, re
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from telegram import (
    Bot, InlineKeyboardButton, InlineKeyboardMarkup,
    ReplyKeyboardMarkup, ReplyKeyboardRemove, Update,
    WebAppInfo,
)
from telegram.ext import (
    Application, CallbackQueryHandler, CommandHandler,
    ContextTypes, ConversationHandler, MessageHandler, filters,
)

# ═══════════════════════ SOZLAMALAR ═══════════════════
BOT_TOKEN       = "7788300032:AAHVu3v_BxGGBAwLf3WQErHaIeNcOst69mI"
ADMIN_IDS       = [7413871633]          # admin Telegram IDlari

# O'qituvchilar: {telegram_id: "varaq_nomi"}
# Masalan: o'qituvchi faqat o'z varag'ini ko'radi
TEACHERS = {
    987654321: "Biloliddin",   # Biloliddin o'qituvchisi
    987654322: "Aziz",         # Aziz o'qituvchisi
}
GITHUB_TOKEN = "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"   # nusxalagan token
GITHUB_REPO  = "bilol1ddin/webapp"          # bu to'g'ri
GITHUB_BRANCH = "main"
EXCEL_FILE      = "oquvchilar.xlsx"
USERS_FILE      = "users.json"
DAVOMAT_FILE    = "davomat.json"
ANALITIKA_FILE  = "analitika.json"
REMINDER_HOUR   = 20
REMINDER_MINUTE = 0
JADVAL_VARAQLAR = ["Biloliddin", "Aziz"]
FUZZY_CHEGARA   = 0.55
# Web App URL — webapp.html ni GitHub Pages yoki boshqa hostga yuklab, URL ni yozing
# Bo'sh qoldirsangiz tugma ko'rinmaydi
WEBAPP_URL      = "https://bilol1ddin.github.io/webapp/"
# ═════════════════════════════════════════════════════


logging.basicConfig(
    format="%(asctime)s - %(levelname)s - %(message)s",
    level=logging.INFO,
    handlers=[logging.FileHandler("bot.log"), logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

ISM_KUTISH, TASDIQLASH, ELON_MATN, ELON_TASDIQ = range(4)

KUN_NOMLARI = {
    0:"Dushanba", 1:"Seshanba", 2:"Chorshanba",
    3:"Payshanba", 4:"Juma",    5:"Shanba",   6:"Yakshanba",
}
KUN_QISQA = {0:"Du", 1:"Se", 2:"Cho", 3:"Pa", 4:"Ju", 5:"Sha"}
SKIP_FANLAR = {"eski","otmen","rad etil","bekor","arxiv"}


# ── Rol tekshiruvlari ─────────────────────────────────
# ══════════════════════════════════════════════════════
# GITHUB AUTO-SYNC (ixtiyoriy)
# ══════════════════════════════════════════════════════

import base64
import urllib.request
import urllib.error

def github_push(filename: str):
    """JSON faylni GitHub ga push qiladi (agar token sozlangan bo'lsa)."""
    if not GITHUB_TOKEN or not GITHUB_REPO:
        return
    try:
        p = Path(filename)
        if not p.exists():
            return
        content = base64.b64encode(p.read_bytes()).decode()
        api_url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{filename}"

        # Mavjud faylning SHA sini olish
        req = urllib.request.Request(api_url)
        req.add_header("Authorization", f"token {GITHUB_TOKEN}")
        req.add_header("Accept", "application/vnd.github.v3+json")
        try:
            with urllib.request.urlopen(req) as resp:
                sha = json.loads(resp.read())["sha"]
        except urllib.error.HTTPError:
            sha = None

        # Push qilish
        data = json.dumps({
            "message": f"Bot: {filename} yangilandi",
            "content": content,
            "branch": GITHUB_BRANCH,
            **({"sha": sha} if sha else {}),
        }).encode()

        req2 = urllib.request.Request(api_url, data=data, method="PUT")
        req2.add_header("Authorization", f"token {GITHUB_TOKEN}")
        req2.add_header("Content-Type", "application/json")
        req2.add_header("Accept", "application/vnd.github.v3+json")
        with urllib.request.urlopen(req2):
            logger.info(f"GitHub sync: {filename}")
    except Exception as e:
        logger.error(f"GitHub sync xato ({filename}): {e}")


def is_admin(uid: int) -> bool:
    return uid in ADMIN_IDS

def is_teacher(uid: int) -> bool:
    return uid in TEACHERS

def teacher_varaq(uid: int) -> str:
    return TEACHERS.get(uid, "")

def is_admin_or_teacher(uid: int) -> bool:
    return is_admin(uid) or is_teacher(uid)


# ══════════════════════════════════════════════════════
# 1. EXCEL
# ══════════════════════════════════════════════════════

def sarlavha_parse(c0, c1):
    t = (str(c0)+str(c1)).upper().replace(" ","")
    if "D/CH/J" in t or "D/S/J" in t:
        kunlar = [0, 2, 4]
    elif "S/P/SH" in t or "S/P/S" in t:
        kunlar = [1, 3, 5]
    else:
        return None
    m = re.search(r"(\d{1,2}:\d{2})", str(c0)+str(c1))
    vaqt = m.group(1) if m else "?"
    om = re.search(r"[\d:]+\s*[/\s]+\s*(.+)", str(c1))
    otuvchi = om.group(1).strip() if om else "—"
    if not otuvchi or len(otuvchi) < 2: otuvchi = "—"
    return {"kunlar": kunlar, "vaqt": vaqt, "otuvchi": otuvchi}


def fn(fan: str) -> str:
    f = fan.strip()
    return f[0].upper()+f[1:].lower() if len(f)>1 else f.upper()


def excel_parse_all(faqat_varaq: str = None) -> list:
    """
    faqat_varaq: faqat shu varaqni o'qiydi (o'qituvchi uchun)
    None bo'lsa — barcha JADVAL_VARAQLAR
    """
    p = Path(EXCEL_FILE)
    if not p.exists():
        logger.error(f"Excel topilmadi: {EXCEL_FILE}")
        return []
    wb = openpyxl.load_workbook(p)
    varaqlar = [faqat_varaq] if faqat_varaq else JADVAL_VARAQLAR
    natija = []
    for vn in varaqlar:
        if vn not in wb.sheetnames:
            continue
        ws = wb[vn]
        joriy = None
        for row in ws.iter_rows(values_only=True):
            c0 = row[0] if len(row)>0 else None
            c1 = row[1] if len(row)>1 else None
            c2 = row[2] if len(row)>2 else None
            g = sarlavha_parse(c0, c1)
            if g: joriy = g; continue
            if not isinstance(c0, int) or not c1 or not joriy: continue
            ism = str(c1).strip()
            fan = str(c2).strip() if c2 else ""
            if not fan or any(s in fan.lower() for s in SKIP_FANLAR): continue
            natija.append({
                "ism": ism, "fan": fn(fan),
                "vaqt": joriy["vaqt"], "kunlar": joriy["kunlar"],
                "otuvchi": joriy["otuvchi"], "varaq": vn,
            })
    wb.close()
    return natija


def excel_search(kiritilgan: str, faqat_varaq: str = None) -> list:
    s = kiritilgan.strip().lower()
    barcha = excel_parse_all(faqat_varaq)
    ism_map: dict[str, list] = {}
    for e in barcha:
        ism_map.setdefault(e["ism"], []).append(e)
    balllar = []
    for ism in ism_map:
        il = ism.lower()
        if s == il or s in il:
            balllar.append((1.0, ism))
        else:
            ball = difflib.SequenceMatcher(None, s, il).ratio()
            il_words = il.split()
            s_words  = s.split()
            if il_words and s_words and il_words[0] == s_words[0]:
                ball = max(ball, 0.72)
            if ball >= FUZZY_CHEGARA:
                balllar.append((ball, ism))
    balllar.sort(key=lambda x: -x[0])
    natija = []
    for _, ism in balllar[:5]:
        natija.extend(ism_map[ism])
    return natija


# ══════════════════════════════════════════════════════
# 2. JSON FAYLLAR
# ══════════════════════════════════════════════════════

def jload(path, default):
    p = Path(path)
    if not p.exists(): return default
    try: return json.loads(p.read_text(encoding="utf-8"))
    except: return default

def jsave(path, data):
    Path(path).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    # GitHub ga avtomatik yuklash
    try:
        github_push(str(path))
    except Exception:
        pass

def users_load(): return jload(USERS_FILE, {})
def users_save(d): jsave(USERS_FILE, d)
def user_set(tid, entries):
    d = users_load(); d[str(tid)] = entries; users_save(d)
def user_get(tid) -> list:
    return users_load().get(str(tid), [])
def user_delete(tid):
    d = users_load(); d.pop(str(tid), None); users_save(d)

def dav_load(): return jload(DAVOMAT_FILE, {"oquvchilar": {}})
def dav_save(d): jsave(DAVOMAT_FILE, d)

def dav_sync(tid: int, ism: str):
    d = dav_load(); sid = str(tid)
    if sid not in d["oquvchilar"]:
        d["oquvchilar"][sid] = {"ism": ism, "davomat": {}}
    else:
        d["oquvchilar"][sid]["ism"] = ism
    dav_save(d)

def dav_belgilay(tid: int, sana: str, keldi: bool):
    d = dav_load(); sid = str(tid)
    if sid not in d["oquvchilar"]:
        # Auto yaratish
        users = users_load()
        ism = users.get(sid, [{}])[0].get("ism", "Noma'lum") if users.get(sid) else "Noma'lum"
        d["oquvchilar"][sid] = {"ism": ism, "davomat": {}}
    d["oquvchilar"][sid]["davomat"][sana] = keldi
    dav_save(d)

def ana_load(): return jload(ANALITIKA_FILE, {"eslatmalar": []})
def ana_save(d): jsave(ANALITIKA_FILE, d)
def ana_qosh(sana, jami, yuborildi, xato):
    d = ana_load()
    d["eslatmalar"].append({
        "sana": sana, "jami": jami,
        "yuborildi": yuborildi, "xato": xato,
        "vaqt": datetime.now().isoformat()
    })
    d["eslatmalar"] = d["eslatmalar"][-90:]
    ana_save(d)


# ══════════════════════════════════════════════════════
# 3. DAVOMAT HISOBLASH
# ══════════════════════════════════════════════════════

def oquvchi_kunlari(tid: int) -> list:
    kunlar = set()
    for e in user_get(tid):
        for k in e.get("kunlar", []):
            if k != 6: kunlar.add(k)
    return sorted(kunlar) or [0,1,2,3,4,5]

def oy_dars_soni(yil, oy, kunlar) -> int:
    _, oxir = calendar.monthrange(yil, oy)
    return sum(1 for k in range(1, oxir+1)
               if date(yil, oy, k).weekday() in kunlar
               and date(yil, oy, k).weekday() != 6)

def oquvchi_stat(tid: int, yil: int, oy: int) -> dict:
    oquvchi = dav_load()["oquvchilar"].get(str(tid))
    if not oquvchi: return None
    oy_str = f"{yil}-{oy:02d}"
    oy_dav = {s:h for s,h in oquvchi["davomat"].items() if s.startswith(oy_str)}
    kunlar = oquvchi_kunlari(tid)
    jami   = oy_dars_soni(yil, oy, kunlar)
    keldi  = sum(1 for v in oy_dav.values() if v)
    kelmadi= sum(1 for v in oy_dav.values() if not v)
    foiz   = round(keldi/jami*100, 1) if jami>0 else 0
    return {
        "ism": oquvchi["ism"], "keldi": keldi, "kelmadi": kelmadi,
        "jami_dars": jami, "foiz": foiz,
        "to_liq": keldi==jami and jami>0, "oy_davomat": oy_dav,
    }

def pbar(foiz: float, n=10) -> str:
    t = round(foiz/100*n)
    return "█"*t + "░"*(n-t)

def foiz_emoji(f):
    if f==100: return "🏆"
    if f>=90:  return "🔥"
    if f>=75:  return "💛"
    if f>=50:  return "⚡"
    if f>0:    return "🔴"
    return "⏳"

def motivatsion(foiz, orni, jami, qolgan):
    if foiz==100: return "🏆 *MUKAMMAL DAVOMAT!*\nNoutbuk o'yiniga qatnashish huquqingiz bor! 🎯"
    if foiz>=90: return f"🔥 *Zo'r!* {orni}-o'rin. Yana *{qolgan}* darsga keling — sovg'ada bo'lasiz! 💪"
    if foiz>=75: return f"💛 *Yaxshi!* {orni}-o'rin. *{qolgan}* dars qoldi — 100% mumkin! 📈"
    if foiz>=50: return f"⚡ *O'rtacha.* {orni}-o'rin ({jami} dan). *{qolgan}* darsni o'tkazib yubormang! 🎁"
    if foiz>0: return f"🔴 *Past.* {orni}-o'rin ({jami} dan). *{qolgan}* darsni o'tkazmang! 🚀"
    return "⏳ *Davomat kiritilmagan.*\nBirinchi darsingizga keling! 🎯"

def reyting_hisobla():
    bugun = date.today()
    natija = []
    for tid in dav_load()["oquvchilar"]:
        s = oquvchi_stat(int(tid), bugun.year, bugun.month)
        if s: natija.append({
            "tg_id": tid, "ism": s["ism"], "foiz": s["foiz"],
            "keldi": s["keldi"], "jami_dars": s["jami_dars"], "to_liq": s["to_liq"],
        })
    natija.sort(key=lambda x: (-x["foiz"], x["ism"]))
    return natija

def barcha_to_liq(yil, oy):
    return [
        {"tg_id": tid, "ism": info["ism"]}
        for tid, info in dav_load()["oquvchilar"].items()
        if (s := oquvchi_stat(int(tid), yil, oy)) and s["to_liq"]
    ]


# ══════════════════════════════════════════════════════
# 4. GURUH TIZIMI — TUZATILGAN
# ══════════════════════════════════════════════════════

def gid_yasatish(otuvchi: str, vaqt: str, kunlar: list) -> str:
    """Xavfsiz guruh IDsi: faqat harf+raqam, max 20 belgi."""
    tur = "D" if 0 in kunlar else "S"
    ot  = re.sub(r"[^A-Za-z0-9]", "", otuvchi)[:8]
    vt  = vaqt.replace(":", "")
    return f"{tur}{vt}{ot}"


def guruhlar_royxati(faqat_varaq: str = None) -> list:
    """
    users.json dagi BARCHA foydalanuvchilarni guruhlariga ajratadi.
    davomat.json da bo'lmasa ham ko'rsatadi (auto sync qiladi).
    faqat_varaq: o'qituvchi uchun faqat o'z guruhlari
    """
    users = users_load()
    guruh_map: dict[str, dict] = {}

    for tg_id, entries in users.items():
        # davomat.json ga auto qo'shish
        dav_d = dav_load()
        if tg_id not in dav_d["oquvchilar"]:
            ism = entries[0]["ism"] if entries else "Noma'lum"
            dav_d["oquvchilar"][tg_id] = {"ism": ism, "davomat": {}}
            dav_save(dav_d)

        ism = dav_load()["oquvchilar"][tg_id]["ism"]

        for e in entries:
            # O'qituvchi filtrasi
            if faqat_varaq and e.get("varaq") != faqat_varaq:
                continue

            gid = gid_yasatish(e.get("otuvchi", ""), e["vaqt"], e["kunlar"])

            if gid not in guruh_map:
                tur = "D/CH/J" if 0 in e["kunlar"] else "S/P/SH"
                guruh_map[gid] = {
                    "id": gid,
                    "otuvchi": e.get("otuvchi", "—"),
                    "vaqt": e["vaqt"],
                    "tur": tur,
                    "kunlar": e["kunlar"],
                    "oquvchilar": [],
                }

            if tg_id not in [o["tg_id"] for o in guruh_map[gid]["oquvchilar"]]:
                guruh_map[gid]["oquvchilar"].append({"tg_id": tg_id, "ism": ism})

    return sorted(guruh_map.values(), key=lambda g: (g["vaqt"], g["otuvchi"]))


def guruh_by_id(gid: str, faqat_varaq: str = None) -> dict:
    return next((g for g in guruhlar_royxati(faqat_varaq) if g["id"] == gid), None)


# ══════════════════════════════════════════════════════
# 5. XABAR MATNLARI
# ══════════════════════════════════════════════════════

def jadval_satrlari(entries):
    if not entries: return "Dars jadvali topilmadi."
    lines = []
    for e in entries:
        kunlar = " · ".join(KUN_QISQA[k] for k in sorted(e["kunlar"]) if k in KUN_QISQA)
        lines.append(
            f"  📖 *{e['fan']}*  🕐 {e['vaqt']}\n"
            f"       👨‍🏫 {e.get('otuvchi','—')}  📅 {kunlar}"
        )
    return "\n".join(lines)

def eslatma_xabari(ism, darslar):
    kun = darslar[0]["kun"]
    lines = [f"📚 *Assalomu alaykum, {ism}!*", "",
             f"Ertaga *{kun}* kuni sizda dars bor:", ""]
    for d in sorted(darslar, key=lambda x: x["vaqt"]):
        lines.append(f"  🕐 *{d['vaqt']}* — {d['fan']}  _(👨‍🏫 {d.get('otuvchi','')})_")
    lines += ["", "Darsga tayor bo'ling! Omad! 🎯"]
    return "\n".join(lines)

def barcha_ertangi_eslatmalar(target):
    if target.weekday() == 6: return {}
    kun_nomi = KUN_NOMLARI[target.weekday()]
    result = {}
    for tg_id, entries in users_load().items():
        darslar = [
            {"ism": e["ism"], "fan": e["fan"], "vaqt": e["vaqt"],
             "kun": kun_nomi, "otuvchi": e.get("otuvchi", "")}
            for e in entries if target.weekday() in e["kunlar"]
        ]
        if darslar: result[tg_id] = darslar
    return result


# ══════════════════════════════════════════════════════
# 6. O'QUVCHI HANDLERLAR
# ══════════════════════════════════════════════════════

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tg_id = update.effective_user.id

    if is_admin(tg_id):
        # Web App tugmasi (agar WEBAPP_URL sozlangan bo'lsa)
        keyboard = []
        if WEBAPP_URL:
            keyboard = [[InlineKeyboardButton(
                "🖥 IT HOUSE Dashboard",
                web_app=WebAppInfo(url=WEBAPP_URL)
            )]]
        await update.message.reply_text(
            f"👑 *IT HOUSE Admin Panel*\n🆔 ID: `{tg_id}`\n\n"
            "*/belgilay* — davomat kiritish\n"
            "*/stat* — statistika\n"
            "*/analitika* — xabar tahlili\n"
            "*/elon* — e'lon yuborish\n"
            "*/loteriya* — sovg'a o'yini\n"
            "*/test* — eslatmani sinash\n"
            "*/royxat* — o'quvchilar ro'yxati\n"
            "*/reload* — Excel yangilash",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard) if keyboard else None,
        )
        return ConversationHandler.END

    if is_teacher(tg_id):
        varaq = teacher_varaq(tg_id)
        await update.message.reply_text(
            f"👨‍🏫 *O'qituvchi Panel*\n🆔 ID: `{tg_id}`\n📚 Varaq: *{varaq}*\n\n"
            "*/belgilay* — o'z guruhingiz davomati\n"
            "*/stat* — guruh statistikasi",
            parse_mode="Markdown",
        )
        return ConversationHandler.END

    mavjud = user_get(tg_id)
    if mavjud:
        ism = mavjud[0]["ism"]
        await update.message.reply_text(
            f"Salom, *{ism}*! 👋\n\n"
            f"📋 *Dars jadvalingiz:*\n{jadval_satrlari(mavjud)}\n\n"
            "/davomat — davomatim\n"
            "/sovga — sovg'a o'yini\n"
            "/reset — qayta ro'yxat",
            parse_mode="Markdown",
        )
        return ConversationHandler.END

    await update.message.reply_text(
        f"🏫 *IT HOUSE Botiga xush kelibsiz!*\n🆔 ID: `{tg_id}`\n\n"
        "Dars eslatmalari uchun\n*to'liq ismingizni* yozing.\n\n"
        "_Masalan:_ `AZIZ KARIMOV`",
        parse_mode="Markdown",
    )
    return ISM_KUTISH


async def ism_qabul(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kiritilgan = update.message.text.strip()
    topilganlar = excel_search(kiritilgan)
    if not topilganlar:
        await update.message.reply_text(
            f"❌ *'{kiritilgan}'* topilmadi.\n\nTo'liq ism-familyangizni yozing.\n_Masalan:_ `AZIZ KARIMOV`",
            parse_mode="Markdown",
        )
        return ISM_KUTISH

    ismlar = list(dict.fromkeys(e["ism"] for e in topilganlar))
    context.user_data["ism_royxat"] = ismlar
    context.user_data["barcha_entries"] = topilganlar

    if len(ismlar) == 1:
        context.user_data["tanlangan_ism"] = ismlar[0]
        return await _tasdiq(update, context)

    tugmalar = [[f"👤 {ism}"] for ism in ismlar[:5]] + [["❌ Hech biri emas"]]
    await update.message.reply_text(
        "Bir nechta topildi, qaysi siz?",
        reply_markup=ReplyKeyboardMarkup(tugmalar, one_time_keyboard=True, resize_keyboard=True),
    )
    return TASDIQLASH


async def _tasdiq(update, context):
    ism = context.user_data.get("tanlangan_ism") or context.user_data["ism_royxat"][0]
    entries = [e for e in context.user_data["barcha_entries"] if e["ism"] == ism]
    context.user_data["tanlangan_entries"] = entries
    tugmalar = ReplyKeyboardMarkup(
        [["✅ Ha, men shuman"], ["❌ Yo'q, boshqa ism"]],
        one_time_keyboard=True, resize_keyboard=True,
    )
    await update.message.reply_text(
        f"Topildi:\n\n👤 *{ism}*\n\n{jadval_satrlari(entries)}\n\nBu sizmisiz?",
        parse_mode="Markdown", reply_markup=tugmalar,
    )
    return TASDIQLASH


async def tasdiqlash(update: Update, context: ContextTypes.DEFAULT_TYPE):
    javob = update.message.text.strip()
    ism_royxat = context.user_data.get("ism_royxat", [])

    if len(ism_royxat) > 1 and javob not in ["✅ Ha, men shuman","❌ Yo'q, boshqa ism","❌ Hech biri emas"]:
        t = javob.replace("👤 ","")
        if t in ism_royxat:
            context.user_data["tanlangan_ism"] = t
            return await _tasdiq(update, context)

    if "Hech biri" in javob or "Yo'q" in javob:
        await update.message.reply_text("Qaytadan kiriting:", reply_markup=ReplyKeyboardRemove())
        return ISM_KUTISH

    if "Ha" in javob:
        tg_id = update.effective_user.id
        entries = context.user_data.get("tanlangan_entries") or [
            e for e in context.user_data["barcha_entries"]
            if e["ism"] == context.user_data["ism_royxat"][0]
        ]
        ism = entries[0]["ism"]
        user_set(tg_id, entries)
        dav_sync(tg_id, ism)
        await update.message.reply_text(
            f"✅ *{ism}, ro'yxatdan o'tdingiz!*\n\n"
            f"Har kuni kechki {REMINDER_HOUR}:00 da eslatma olasiz.\n\n"
            f"📋 *Jadvalingiz:*\n{jadval_satrlari(entries)}\n\n/davomat",
            parse_mode="Markdown", reply_markup=ReplyKeyboardRemove(),
        )
        try:
            await context.bot.send_message(
                chat_id=ADMIN_IDS[0],
                text=f"🆕 Yangi o'quvchi:\n👤 {ism}\nID: `{tg_id}`",
                parse_mode="Markdown",
            )
        except Exception: pass
        return ConversationHandler.END

    await update.message.reply_text("Qaytadan kiriting:", reply_markup=ReplyKeyboardRemove())
    return ISM_KUTISH


async def reset_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_delete(update.effective_user.id)
    await update.message.reply_text("♻️ O'chirildi. /start", reply_markup=ReplyKeyboardRemove())
    return ConversationHandler.END

async def bekor(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Bekor.", reply_markup=ReplyKeyboardRemove())
    return ConversationHandler.END


# ── /davomat ──────────────────────────────────────────

async def davomat_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tg_id = update.effective_user.id
    bugun = date.today()
    entries = user_get(tg_id)
    if entries: dav_sync(tg_id, entries[0]["ism"])
    stat = oquvchi_stat(tg_id, bugun.year, bugun.month)
    if not stat:
        await update.message.reply_text("Hali ro'yxatdan o'tmagansiz.\n/start ni bosing.")
        return

    oy_nomi = bugun.strftime("%B %Y")
    _, oxir = calendar.monthrange(bugun.year, bugun.month)
    kunlar_list = oquvchi_kunlari(tg_id)
    oy_dav = stat["oy_davomat"]

    # Kalentar ko'rinish
    kun_qatorlari = []
    for k in range(1, oxir+1):
        d = date(bugun.year, bugun.month, k)
        if d.weekday() not in kunlar_list or d.weekday() == 6: continue
        sana_str = d.strftime("%Y-%m-%d")
        holat = oy_dav.get(sana_str)
        if holat is True: belgi = "✅"
        elif holat is False: belgi = "❌"
        elif d > bugun: belgi = "🔲"
        else: belgi = "⬜"
        kun_qatorlari.append(f"{belgi}{d.day:02d}")

    kal_qatorlar = []
    for i in range(0, len(kun_qatorlari), 5):
        kal_qatorlar.append("  "+"  ".join(kun_qatorlari[i:i+5]))
    kal_str = "\n".join(kal_qatorlar) or "  _Dars kunlari yo'q_"

    bar = pbar(stat["foiz"], 12)
    emoji = foiz_emoji(stat["foiz"])
    reyting = reyting_hisobla()
    jami_oquvchi = len(reyting)
    ornim = next((i+1 for i,r in enumerate(reyting) if r["tg_id"]==str(tg_id)), jami_oquvchi)
    qolgan = max(0, stat["jami_dars"]-stat["keldi"])

    medal = {1:"🥇", 2:"🥈", 3:"🥉"}
    top = []
    for i, r in enumerate(reyting[:5], 1):
        men = " ◀ *Siz*" if r["tg_id"]==str(tg_id) else ""
        top.append(f"{medal.get(i,f'  {i}.')} `{pbar(r['foiz'],8)}` *{r['foiz']}%* {r['ism'][:16]}{men}")
    if ornim > 5:
        m = next((r for r in reyting if r["tg_id"]==str(tg_id)), None)
        if m:
            top.append("  `┄┄┄┄┄┄┄┄`")
            top.append(f"  {ornim}. `{pbar(m['foiz'],8)}` *{m['foiz']}%* {m['ism'][:16]} ◀ *Siz*")

    motiv = motivatsion(stat["foiz"], ornim, jami_oquvchi, qolgan)

    await update.message.reply_text(
        f"┌─────────────────────┐\n"
        f"│   📊 DAVOMAT        │\n"
        f"└─────────────────────┘\n\n"
        f"👤 *{stat['ism']}*  |  {oy_nomi}\n\n"
        f"─────────────────────\n"
        f"{kal_str}\n\n"
        f"✅ keldi  ❌ kelmadi  ⬜ o'tdi  🔲 kelasi\n\n"
        f"─────────────────────\n"
        f"{emoji} `{bar}` *{stat['foiz']}%*\n\n"
        f"  ✅ Keldi:     *{stat['keldi']}* kun\n"
        f"  ❌ Kelmadi:   *{stat['kelmadi']}* kun\n"
        f"  📚 Jami:      *{stat['jami_dars']}* kun\n\n"
        f"─────────────────────\n"
        f"🏆 *Reyting: {ornim}/{jami_oquvchi}*\n\n"
        + "\n".join(top) +
        f"\n\n─────────────────────\n{motiv}",
        parse_mode="Markdown",
    )


async def sovga_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    bugun = date.today()
    _, oxirgi = calendar.monthrange(bugun.year, bugun.month)
    qolgan_kun = (date(bugun.year, bugun.month, oxirgi) - bugun).days
    tg_id = update.effective_user.id
    stat = oquvchi_stat(tg_id, bugun.year, bugun.month)
    to_liqlar = barcha_to_liq(bugun.year, bugun.month)
    n_dav = len(dav_load()["oquvchilar"])

    if stat and stat["to_liq"]:
        shaxsiy = "🏆 *Siz 100%! O'yinda qatnashish huquqingiz bor!*"
    elif stat and stat["foiz"] > 0:
        shaxsiy = f"💡 Yana *{stat['jami_dars']-stat['keldi']}* darsga keling — o'yiniga tusharsiz!"
    else:
        shaxsiy = "_Davomat hali kiritilmagan._"

    await update.message.reply_text(
        "┌─────────────────────┐\n"
        "│  🎁 SOVG'A O'YINI  │\n"
        "└─────────────────────┘\n\n"
        "📜 *Shartlar:*\n"
        "  • Oyda barcha darslarga keling (100%)\n"
        "  • Yakshanba — dam olish (hisoblanmaydi)\n"
        "  • Oy oxirida tasodifiy 1 ta tanlanadi\n\n"
        "💻 *Katta sovg'a: NOUTBUK!*\n\n"
        f"⏳ Oy tugashiga: *{qolgan_kun} kun*\n"
        f"🏆 100% davomat: *{len(to_liqlar)}/{n_dav}* o'quvchi\n"
        f"`{pbar(len(to_liqlar)/max(n_dav,1)*100, 12)}`\n\n"
        f"─────────────────────\n{shaxsiy}",
        parse_mode="Markdown",
    )


# ══════════════════════════════════════════════════════
# 7. DAVOMAT BELGILASH — TUZATILGAN
# ══════════════════════════════════════════════════════
# Callback format:
#   G:{gid}              — guruh tanlash
#   GS:{gid}:{sana}      — sana tanlash
#   GT:{tid}:{sana}:{gid} — toggle
#   GA:{tf}:{sana}:{gid}  — all true/false
#   GBACK                — orqaga
# ══════════════════════════════════════════════════════

async def belgilay_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin_or_teacher(update.effective_user.id):
        await update.message.reply_text("⛔ Faqat admin yoki o'qituvchilar uchun.")
        return

    uid = update.effective_user.id
    faqat_varaq = teacher_varaq(uid) if is_teacher(uid) and not is_admin(uid) else None

    guruhlar = guruhlar_royxati(faqat_varaq)
    if not guruhlar:
        await update.message.reply_text(
            "Guruhlar topilmadi.\n\n"
            "O'quvchilar /start bosib ro'yxatdan o'tishi kerak!"
        )
        return

    buttons = []
    for g in guruhlar:
        n = len(g["oquvchilar"])
        label = f"🕐{g['vaqt']} 👨‍🏫{g['otuvchi'][:10]} ({g['tur']}) — {n} o'q."
        buttons.append([InlineKeyboardButton(label, callback_data=f"G:{g['id']}")])

    await update.message.reply_text(
        "📋 *Guruhni tanlang:*\n_(vaqt · o'qituvchi · kun turi)_",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def cb_guruh(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    gid = query.data.split(":", 1)[1]

    uid = query.from_user.id
    faqat_varaq = teacher_varaq(uid) if is_teacher(uid) and not is_admin(uid) else None
    guruh = guruh_by_id(gid, faqat_varaq)

    if not guruh:
        await query.edit_message_text("Guruh topilmadi. /belgilay ni qaytadan bosing.")
        return

    # Oxirgi 14 kundan faqat dars kunlari
    buttons = []
    for i in range(14):
        kun = date.today() - timedelta(days=i)
        if kun.weekday() not in guruh["kunlar"] or kun.weekday() == 6:
            continue
        sana_str = kun.strftime("%Y-%m-%d")
        buttons.append([InlineKeyboardButton(
            f"📅 {kun.strftime('%d-%b, %A')}",
            callback_data=f"GS:{gid}:{sana_str}"
        )])
    buttons.append([InlineKeyboardButton("🔙 Orqaga", callback_data="GBACK")])

    tur_str = "Du·Cho·Ju" if 0 in guruh["kunlar"] else "Se·Pa·Sha"
    await query.edit_message_text(
        f"👨‍🏫 *{guruh['otuvchi']}*  🕐{guruh['vaqt']}  ({tur_str})\n"
        f"👥 {len(guruh['oquvchilar'])} o'quvchi\n\n"
        "📅 Dars sanasini tanlang:",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def cb_guruh_sana(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    # GS:{gid}:{sana}
    _, gid, sana_str = query.data.split(":", 2)
    await _davomat_sahifa(query, gid, sana_str)


async def cb_toggle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    # GT:{tid}:{sana}:{gid}
    parts = query.data.split(":")
    tid_str, sana_str, gid = parts[1], parts[2], parts[3]

    dav = dav_load()
    mavjud = dav["oquvchilar"].get(tid_str, {}).get("davomat", {}).get(sana_str)
    yangi = False if mavjud is True else True
    dav_belgilay(int(tid_str), sana_str, yangi)
    await _davomat_sahifa(query, gid, sana_str)


async def cb_all(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    # GA:{true/false}:{sana}:{gid}
    parts = query.data.split(":")
    tf, sana_str, gid = parts[1], parts[2], parts[3]
    keldi = tf == "true"

    uid = query.from_user.id
    faqat_varaq = teacher_varaq(uid) if is_teacher(uid) and not is_admin(uid) else None
    guruh = guruh_by_id(gid, faqat_varaq)
    if guruh:
        for oq in guruh["oquvchilar"]:
            dav_belgilay(int(oq["tg_id"]), sana_str, keldi)
    await _davomat_sahifa(query, gid, sana_str)


async def cb_gback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    uid = query.from_user.id
    faqat_varaq = teacher_varaq(uid) if is_teacher(uid) and not is_admin(uid) else None
    guruhlar = guruhlar_royxati(faqat_varaq)
    buttons = []
    for g in guruhlar:
        n = len(g["oquvchilar"])
        label = f"🕐{g['vaqt']} 👨‍🏫{g['otuvchi'][:10]} ({g['tur']}) — {n} o'q."
        buttons.append([InlineKeyboardButton(label, callback_data=f"G:{g['id']}")])
    await query.edit_message_text(
        "📋 *Guruhni tanlang:*",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def _davomat_sahifa(query, gid: str, sana_str: str):
    """Guruh davomat sahifasini ko'rsatadi."""
    uid = query.from_user.id
    faqat_varaq = teacher_varaq(uid) if is_teacher(uid) and not is_admin(uid) else None
    guruh = guruh_by_id(gid, faqat_varaq)

    if not guruh:
        await query.edit_message_text("Xato. /belgilay ni qaytadan bosing.")
        return

    d_obj = datetime.strptime(sana_str, "%Y-%m-%d")
    dav = dav_load()["oquvchilar"]
    keldi_n = kelmadi_n = belgisiz_n = 0
    buttons = []

    for oq in guruh["oquvchilar"]:
        mavjud = dav.get(str(oq["tg_id"]), {}).get("davomat", {}).get(sana_str)
        if mavjud is True:
            belgi = "✅"; keldi_n += 1
        elif mavjud is False:
            belgi = "❌"; kelmadi_n += 1
        else:
            belgi = "⬜"; belgisiz_n += 1
        buttons.append([InlineKeyboardButton(
            f"{belgi} {oq['ism'][:30]}",
            callback_data=f"GT:{oq['tg_id']}:{sana_str}:{gid}"
        )])

    # Tezkor tugmalar
    tezkor = []
    if keldi_n < len(guruh["oquvchilar"]):
        tezkor.append(InlineKeyboardButton("✅ Barchasi keldi", callback_data=f"GA:true:{sana_str}:{gid}"))
    if kelmadi_n < len(guruh["oquvchilar"]):
        tezkor.append(InlineKeyboardButton("❌ Hech kim kelmadi", callback_data=f"GA:false:{sana_str}:{gid}"))
    if tezkor:
        buttons.insert(0, tezkor)

    buttons.append([InlineKeyboardButton("🔙 Orqaga", callback_data=f"G:{gid}")])

    n = len(guruh["oquvchilar"])
    foiz_g = round(keldi_n/n*100) if n else 0

    await query.edit_message_text(
        f"👨‍🏫 *{guruh['otuvchi']}*  |  📅 *{d_obj.strftime('%d %B, %A')}*\n\n"
        f"`{pbar(foiz_g,12)}` *{foiz_g}%*\n"
        f"✅ {keldi_n}  ❌ {kelmadi_n}  ⬜ {belgisiz_n}\n\n"
        f"Bosing — holat almashadi:",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(buttons),
    )


# ══════════════════════════════════════════════════════
# 8. ADMIN BUYRUQLARI
# ══════════════════════════════════════════════════════

async def stat_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin_or_teacher(update.effective_user.id): return
    uid = update.effective_user.id
    faqat_varaq = teacher_varaq(uid) if is_teacher(uid) and not is_admin(uid) else None
    bugun = date.today()
    guruhlar = guruhlar_royxati(faqat_varaq)
    if not guruhlar:
        await update.message.reply_text("Ma'lumot yo'q. O'quvchilar ro'yxatdan o'tishi kerak.")
        return

    oy_nomi = bugun.strftime("%B %Y")
    jami_oquvchi = len(dav_load()["oquvchilar"])
    jami_to_liq  = len(barcha_to_liq(bugun.year, bugun.month))
    lines = [
        f"┌─────────────────────┐\n│ 📊 {oy_nomi}\n└─────────────────────┘\n",
        f"👥 Jami: *{jami_oquvchi}*  🏆 100%: *{jami_to_liq}*\n",
    ]
    for g in guruhlar:
        n = len(g["oquvchilar"])
        if n == 0: continue
        jami_dars = oy_dars_soni(bugun.year, bugun.month, g["kunlar"])
        if jami_dars == 0: continue
        keldi_jami = to_liq_n = 0
        for oq in g["oquvchilar"]:
            s = oquvchi_stat(int(oq["tg_id"]), bugun.year, bugun.month)
            if s:
                keldi_jami += s["keldi"]
                if s["to_liq"]: to_liq_n += 1
        foiz_g = round(keldi_jami/(n*jami_dars)*100, 1) if n*jami_dars > 0 else 0
        tur = "Du·Cho·Ju" if 0 in g["kunlar"] else "Se·Pa·Sha"
        lines.append(
            f"👨‍🏫 *{g['otuvchi']}*  🕐{g['vaqt']}  _{tur}_\n"
            f"  `{pbar(foiz_g,8)}` *{foiz_g}%*  {n} o'q  🏆{to_liq_n}\n"
        )
    matn = "\n".join(lines)
    if len(matn) > 4000: matn = matn[:4000] + "\n..."
    await update.message.reply_text(matn, parse_mode="Markdown")


async def analitika_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id): return
    d = ana_load()
    eslatmalar = d.get("eslatmalar", [])
    if not eslatmalar:
        await update.message.reply_text("Hali analitika yo'q. /test yuboring.")
        return

    oxirgi7 = eslatmalar[-7:]
    jami_y = sum(e["yuborildi"] for e in oxirgi7)
    jami_x = sum(e["xato"] for e in oxirgi7)
    jami_j = sum(e["jami"] for e in oxirgi7)
    foiz_m = round(jami_y/jami_j*100, 1) if jami_j > 0 else 0

    lines = [
        "┌──────────────────────┐\n│  📨 XABAR ANALITIKA  │\n└──────────────────────┘\n",
        f"*Oxirgi 7 ta eslatma:*\n"
        f"  📤 Yuborildi:  *{jami_y}*\n"
        f"  ❌ Xato:       *{jami_x}*\n"
        f"  `{pbar(foiz_m,12)}` *{foiz_m}%*\n\n"
        f"─────────────────────\n*Kunlik:*\n",
    ]
    for e in reversed(oxirgi7):
        d_o = datetime.fromisoformat(e["vaqt"])
        fe = round(e["yuborildi"]/e["jami"]*100) if e["jami"]>0 else 0
        lines.append(
            f"📅 {d_o.strftime('%d-%b')} `{pbar(fe,6)}` {fe}%  ✅{e['yuborildi']} ❌{e['xato']} /{e['jami']}"
        )
    lines.append(f"\n─────────────────────")
    lines.append(f"👥 Faol: *{len(users_load())}*  📊 Davomat: *{len(dav_load()['oquvchilar'])}*")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


async def elon_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Faqat adminlar uchun.")
        return
    await update.message.reply_text(
        "📢 *E'lon matnini yozing:*\n\nBekor: /bekor", parse_mode="Markdown"
    )
    return ELON_MATN


async def elon_matn_qabul(update: Update, context: ContextTypes.DEFAULT_TYPE):
    matn = update.message.text.strip()
    context.user_data["elon_matn"] = matn
    n = len(users_load())
    tugmalar = ReplyKeyboardMarkup([["✅ Ha, yubor"],["❌ Bekor"]], one_time_keyboard=True, resize_keyboard=True)
    await update.message.reply_text(
        f"📢 *Ko'rinish:*\n\n─────────────────────\n"
        f"📣 *IT HOUSE E'LONI*\n\n{matn}\n─────────────────────\n\n"
        f"👥 *{n} ta* o'quvchiga yuboriladi. Tasdiqlaysizmi?",
        parse_mode="Markdown", reply_markup=tugmalar,
    )
    return ELON_TASDIQ


async def elon_tasdiqlash(update: Update, context: ContextTypes.DEFAULT_TYPE):
    javob = update.message.text.strip()
    if "Ha" not in javob:
        await update.message.reply_text("Bekor.", reply_markup=ReplyKeyboardRemove())
        return ConversationHandler.END
    matn = context.user_data.get("elon_matn","")
    elon_xabar = f"📣 *IT HOUSE E'LONI*\n\n{matn}"
    users = users_load()
    await update.message.reply_text(f"🚀 {len(users)} ta o'quvchiga yuborilmoqda...", reply_markup=ReplyKeyboardRemove())
    y = x = 0
    for tg_id in users:
        try:
            await context.bot.send_message(chat_id=int(tg_id), text=elon_xabar, parse_mode="Markdown")
            y += 1
        except Exception as e:
            x += 1; logger.error(f"E'lon xato {tg_id}: {e}")
    await update.message.reply_text(f"✅ Yuborildi: *{y}*  ❌ Xato: *{x}*", parse_mode="Markdown")
    return ConversationHandler.END


async def loteriya_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Faqat adminlar uchun.")
        return
    bugun = date.today()
    to_liqlar = barcha_to_liq(bugun.year, bugun.month)
    if not to_liqlar:
        await update.message.reply_text("⚠️ 100% davomat qilgan o'quvchi yo'q.")
        return
    ismlar = "\n".join(f"  🏆 {t['ism']}" for t in to_liqlar)
    await update.message.reply_text(
        f"🎰 *SOVG'A O'YINI*\n\n100% davomat ({len(to_liqlar)} ta):\n{ismlar}\n\n🎲 Tanlanmoqda...",
        parse_mode="Markdown",
    )
    import asyncio; await asyncio.sleep(3)
    golib = random.choice(to_liqlar)
    try:
        await context.bot.send_message(
            chat_id=int(golib["tg_id"]),
            text=f"🎉 *TABRIKLAYMIZ, {golib['ism']}!*\n\nSiz g'alibsiz!\n💻 Sovg'a: *NOUTBUK*\n\nMarkaz bilan bog'laning! 🏫",
            parse_mode="Markdown",
        )
        xabar = "✅ G'olibga xabar yuborildi!"
    except Exception:
        xabar = "⚠️ G'olibga xabar yuborib bo'lmadi"
    await update.message.reply_text(f"🏆 *G'OLIB:*\n\n👑 *{golib['ism']}*\n\n{xabar}", parse_mode="Markdown")


async def test_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Faqat adminlar uchun.")
        return
    target = datetime.now() + timedelta(days=1)
    if target.weekday() == 6:
        await update.message.reply_text("⚠️ Ertaga Yakshanba — dam olish, eslatma yo'q.")
        return
    darslar_map = barcha_ertangi_eslatmalar(target)
    if not darslar_map:
        await update.message.reply_text(f"Ertaga ({KUN_NOMLARI[target.weekday()]}) eslatma yo'q.")
        return
    await update.message.reply_text(f"🚀 {len(darslar_map)} ta o'quvchiga yuborilmoqda...")
    y = x = 0
    for tg_id, darslar in darslar_map.items():
        try:
            await context.bot.send_message(
                chat_id=int(tg_id), text=eslatma_xabari(darslar[0]["ism"], darslar), parse_mode="Markdown"
            )
            y += 1
        except Exception as e:
            x += 1; await update.message.reply_text(f"⚠️ {tg_id}: {e}")
    ana_qosh(target.strftime("%Y-%m-%d"), len(darslar_map), y, x)
    await update.message.reply_text(f"✅ {y} yuborildi  ❌ {x} xato")


async def royxat_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id): return
    users = users_load()
    if not users:
        await update.message.reply_text("Ro'yxat bo'sh.")
        return
    lines = [f"👥 *Ro'yxat ({len(users)} ta):*\n"]
    for tg_id, entries in users.items():
        ism = entries[0]["ism"] if entries else "?"
        ot  = entries[0].get("otuvchi","—") if entries else "—"
        lines.append(f"  • {ism}  _{ot}_  `{tg_id}`")
    matn = "\n".join(lines)
    if len(matn) > 4000: matn = matn[:4000]+"\n..."
    await update.message.reply_text(matn, parse_mode="Markdown")


async def reload_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id): return
    barcha = excel_parse_all()
    await update.message.reply_text(f"✅ Excel qayta o'qildi! Jami: {len(barcha)}")


# ══════════════════════════════════════════════════════
# 9. KUNLIK ESLATMA
# ══════════════════════════════════════════════════════

async def yuborish_barcha(bot: Bot):
    target = datetime.now() + timedelta(days=1)
    if target.weekday() == 6:
        logger.info("Ertaga Yakshanba — eslatma yo'q.")
        return
    kun_nomi = KUN_NOMLARI[target.weekday()]
    logger.info(f"Eslatma: {kun_nomi} {target.strftime('%d.%m.%Y')}")
    darslar_map = barcha_ertangi_eslatmalar(target)
    y = x = 0
    for tg_id, darslar in darslar_map.items():
        try:
            await bot.send_message(
                chat_id=int(tg_id), text=eslatma_xabari(darslar[0]["ism"], darslar), parse_mode="Markdown"
            )
            y += 1
        except Exception as e:
            x += 1; logger.error(f"XATO {tg_id}: {e}")
    ana_qosh(target.strftime("%Y-%m-%d"), len(darslar_map), y, x)
    try:
        await bot.send_message(
            chat_id=ADMIN_IDS[0],
            text=f"📊 *Kunlik hisobot*\n📅 {kun_nomi}, {target.strftime('%d.%m.%Y')}\n"
                 f"✅ {y}  ❌ {x}  |  {round(y/len(darslar_map)*100) if darslar_map else 0}%",
            parse_mode="Markdown",
        )
    except Exception: pass


async def scheduled_job(context: ContextTypes.DEFAULT_TYPE):
    await yuborish_barcha(context.bot)


# ══════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════

async def post_init(app: Application):
    """Bot ishga tushganda menu va komandalarni o'rnatadi."""
    from telegram import MenuButtonWebApp, BotCommand
    # Admin uchun menu button — faqat admin ko'radi
    # (Telegram hozircha barcha foydalanuvchilarga bir xil menu qo'yadi,
    #  shuning uchun webappni /start da admin uchun ko'rsatamiz)
    await app.bot.set_my_commands([
        BotCommand("start",    "Boshlash / Mening jadvalim"),
        BotCommand("davomat",  "Davomatim va reyting"),
        BotCommand("sovga",    "Sovg'a o'yini"),
        BotCommand("reset",    "Qayta ro'yxatdan o'tish"),
    ])
    logger.info("Bot komandalar o'rnatildi.")


def main():
    app = Application.builder().token(BOT_TOKEN).post_init(post_init).build()

    conv = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            ISM_KUTISH: [MessageHandler(filters.TEXT & ~filters.COMMAND, ism_qabul)],
            TASDIQLASH: [MessageHandler(filters.TEXT & ~filters.COMMAND, tasdiqlash)],
        },
        fallbacks=[CommandHandler("bekor", bekor), CommandHandler("reset", reset_cmd)],
    )
    elon_conv = ConversationHandler(
        entry_points=[CommandHandler("elon", elon_cmd)],
        states={
            ELON_MATN:  [MessageHandler(filters.TEXT & ~filters.COMMAND, elon_matn_qabul)],
            ELON_TASDIQ:[MessageHandler(filters.TEXT & ~filters.COMMAND, elon_tasdiqlash)],
        },
        fallbacks=[CommandHandler("bekor", bekor)],
    )

    app.add_handler(conv)
    app.add_handler(elon_conv)
    app.add_handler(CommandHandler("reset",     reset_cmd))
    app.add_handler(CommandHandler("davomat",   davomat_cmd))
    app.add_handler(CommandHandler("sovga",     sovga_cmd))
    app.add_handler(CommandHandler("belgilay",  belgilay_cmd))
    app.add_handler(CommandHandler("stat",      stat_cmd))
    app.add_handler(CommandHandler("analitika", analitika_cmd))
    app.add_handler(CommandHandler("loteriya",  loteriya_cmd))
    app.add_handler(CommandHandler("test",      test_cmd))
    app.add_handler(CommandHandler("royxat",    royxat_cmd))
    app.add_handler(CommandHandler("reload",    reload_cmd))

    # Callback handlers — yangi format G: GS: GT: GA: GBACK
    app.add_handler(CallbackQueryHandler(cb_guruh,      pattern=r"^G:(?!S:|T:|A:)"))
    app.add_handler(CallbackQueryHandler(cb_guruh_sana, pattern=r"^GS:"))
    app.add_handler(CallbackQueryHandler(cb_toggle,     pattern=r"^GT:"))
    app.add_handler(CallbackQueryHandler(cb_all,        pattern=r"^GA:"))
    app.add_handler(CallbackQueryHandler(cb_gback,      pattern=r"^GBACK$"))

    now = datetime.now()
    first_run = now.replace(hour=REMINDER_HOUR, minute=REMINDER_MINUTE, second=0, microsecond=0)
    if first_run <= now: first_run += timedelta(days=1)

    app.job_queue.run_repeating(
        scheduled_job, interval=timedelta(days=1),
        first=first_run, name="kunlik_eslatma",
    )

    logger.info(f"Bot ishga tushdi. Eslatma: {first_run.strftime('%d.%m.%Y %H:%M')}")
    app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
